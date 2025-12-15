#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
阿里云ECS智能报价系统 - Streamlit前端界面
简洁设计：仅作为用户交互入口,所有业务逻辑由后端处理
"""
import streamlit as st
import pandas as pd
import os
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
import sys
import logging
from io import StringIO

# 添加项目路径到系统路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.pricing_service import PricingService
from app.core.sku_recommend_service import SKURecommendService
from app.data.batch_processor import BatchQuotationProcessor
from app.data.data_ingestion import LLMDrivenExcelLoader

# 加载环境变量
load_dotenv()

# ============================================================================
# 日志系统配置
# ============================================================================

class StreamlitLogHandler(logging.Handler):
    """
    自定义日志处理器 - 将日志输出到Streamlit组件
    """
    def __init__(self):
        super().__init__()
        self.log_buffer = []
        self.max_logs = 200  # 最多保存200条日志
    
    def emit(self, record):
        try:
            msg = self.format(record)
            timestamp = datetime.fromtimestamp(record.created).strftime('%H:%M:%S')
            
            # 根据日志级别添加emoji
            if record.levelno >= logging.ERROR:
                prefix = "❌"
            elif record.levelno >= logging.WARNING:
                prefix = "⚠️"
            elif record.levelno >= logging.INFO:
                prefix = "ℹ️"
            else:
                prefix = "🔍"
            
            formatted_msg = f"[{timestamp}] {prefix} {msg}"
            self.log_buffer.append(formatted_msg)
            
            # 保持日志数量在限制之内
            if len(self.log_buffer) > self.max_logs:
                self.log_buffer.pop(0)
        except Exception:
            self.handleError(record)
    
    def get_logs(self):
        """获取所有日志"""
        return self.log_buffer
    
    def clear_logs(self):
        """清除日志"""
        self.log_buffer.clear()


def setup_logging():
    """
    配置日志系统：同时输出到控制台和Streamlit
    """
    # 创建Streamlit日志处理器
    if 'log_handler' not in st.session_state:
        st.session_state.log_handler = StreamlitLogHandler()
        
        # 配置日志格式
        formatter = logging.Formatter('%(message)s')
        st.session_state.log_handler.setFormatter(formatter)
        
        # 将处理器添加到root logger
        root_logger = logging.getLogger()
        root_logger.setLevel(logging.INFO)
        root_logger.addHandler(st.session_state.log_handler)
        
        # 也添加到app模块的logger
        app_logger = logging.getLogger('app')
        app_logger.setLevel(logging.INFO)
        app_logger.addHandler(st.session_state.log_handler)
    
    return st.session_state.log_handler

# ============================================================================
# 页面配置
# ============================================================================
st.set_page_config(
    page_title="阿里云ECS智能报价系统",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# 简洁样式
# ============================================================================
st.markdown("""
<style>
    .stButton>button {
        background-color: #FF6A00;
        color: white;
    }
    .stButton>button:hover {
        background-color: #E65A00;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# 辅助函数
# ============================================================================

@st.cache_data
def get_region_options():
    """获取阿里云区域选项"""
    return {
        "华北2（北京）": "cn-beijing",
        "华东2（上海）": "cn-shanghai",
        "华东1（杭州）": "cn-hangzhou",
        "华南1（深圳）": "cn-shenzhen",
        "华南2（广州）": "cn-guangzhou",
        "华北1（青岛）": "cn-qingdao",
        "华北3（张家口）": "cn-zhangjiakou",
        "西南1（成都）": "cn-chengdu",
        "香港": "cn-hongkong",
        "亚太东南1（新加坡）": "ap-southeast-1",
        "亚太东南5（雅加达）": "ap-southeast-5",
        "美国西部1（硅谷）": "us-west-1",
        "美国东部1（弗吉尼亚）": "us-east-1",
        "欧洲中部1（法兰克福）": "eu-central-1",
    }


def initialize_services(region_id: str):
    """
    初始化所有服务组件
    
    Args:
        region_id: 阿里云区域ID（如 cn-beijing）
        
    Returns:
        tuple: (pricing_service, sku_service, processor)
    """
    # 获取环境变量
    access_key_id = os.getenv('ALIBABA_CLOUD_ACCESS_KEY_ID')
    access_key_secret = os.getenv('ALIBABA_CLOUD_ACCESS_KEY_SECRET')
    dashscope_key = os.getenv('DASHSCOPE_API_KEY')
    
    if not access_key_id or not access_key_secret:
        st.error("❌ 缺少阿里云API密钥，请检查.env文件配置")
        st.stop()
    
    if not dashscope_key:
        st.warning("⚠️ 缺少DashScope API密钥，AI解析功能将不可用")
    
    # 初始化价格查询服务
    pricing_service = PricingService(
        access_key_id=access_key_id,
        access_key_secret=access_key_secret,
        region_id=region_id
    )
    
    # 初始化SKU推荐服务
    sku_service = SKURecommendService(
        access_key_id=access_key_id,
        access_key_secret=access_key_secret,
        region_id=region_id
    )
    
    # 初始化批量处理器
    processor = BatchQuotationProcessor(
        pricing_service=pricing_service,
        sku_recommend_service=sku_service,
        region=region_id
    )
    
    return pricing_service, sku_service, processor


def save_uploaded_file(uploaded_file):
    """保存上传文件到临时目录"""
    temp_dir = Path("temp_uploads")
    temp_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = temp_dir / f"{timestamp}_{uploaded_file.name}"
    
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    
    return file_path


# ============================================================================
# 侧边栏：配置
# ============================================================================
with st.sidebar:
    st.markdown("### ⚙️ 配置")
    
    # 地域选择
    region_options = get_region_options()
    selected_region = st.selectbox(
        "目标地域",
        options=list(region_options.keys()),
        index=0
    )
    
    region_id = region_options[selected_region]
    st.info(f"区域: **{region_id}**")
    
    st.markdown("---")
    st.caption("💡 计费模式: 包年包月")
    st.caption("🎯 推荐策略: NewProductFirst")
    
    # 日志查看区域
    st.markdown("---")
    st.markdown("### 📜 处理日志")
    
    if 'log_handler' in st.session_state:
        logs = st.session_state.log_handler.get_logs()
        if logs:
            # 显示最后20条日志
            recent_logs = logs[-20:] if len(logs) > 20 else logs
            st.code('\n'.join(recent_logs), language='log', line_numbers=False)
        else:
            st.info("📋 暂无日志")
    else:
        st.info("📋 暂无日志")

# ============================================================================
# 主界面
# ============================================================================

# 初始化日志系统
log_handler = setup_logging()

st.title("💰 阿里云ECS智能报价系统")
st.caption("上传Excel文件，自动生成报价单")

st.markdown("---")

# 文件上传
st.subheader("📁 上传Excel文件")

uploaded_file = st.file_uploader(
    "选择Excel文件",
    type=['xlsx', 'xls'],
    label_visibility="collapsed"
)

if uploaded_file:
    st.success(f"✅ 已选择: {uploaded_file.name}")
    
    # 日志展示区域（位于文件选择后）
    with st.expander("📜 处理日志", expanded=False):
        log_container = st.empty()
        
        # 清除日志按钮
        col_clear1, col_clear2 = st.columns([1, 5])
        with col_clear1:
            if st.button("🗑️ 清除日志"):
                log_handler.clear_logs()
                st.rerun()
    
    # 开始处理按钮
    if st.button("🚀 开始生成报价", type="primary", use_container_width=True):
        start_processing = True
    else:
        start_processing = False
    
    # 处理逻辑
    if start_processing:
        # 清除旧日志
        log_handler.clear_logs()
        
        with st.spinner("⚙️ 正在初始化服务..."):
            try:
                logging.info("🚀 开始初始化服务...")
                
                # 初始化服务
                pricing_service, sku_service, processor = initialize_services(region_id)
                logging.info(f"✅ 服务初始化完成 (区域: {region_id})")
                
                # 保存文件
                file_path = save_uploaded_file(uploaded_file)
                logging.info(f"📁 文件已保存: {file_path.name}")
                
                # 创建LLM加载器并调用后端处理
                logging.info("🤖 使用AI智能解析 (LLMDrivenExcelLoader)")
                
                # 读取所有Sheet
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                all_sheets = wb.sheetnames
                logging.info(f"📋 检测到 {len(all_sheets)} 个工作表: {', '.join(all_sheets)}")
                
                # 更新日志显示
                log_container.code('\n'.join(log_handler.get_logs()), language='log')
                
                all_results = []
                
                # 遍历处理每个Sheet
                for sheet_idx, sheet_name in enumerate(all_sheets, 1):
                    logging.info(f"\n{'='*60}")
                    logging.info(f"🔄 处理工作表 [{sheet_idx}/{len(all_sheets)}]: {sheet_name}")
                    logging.info(f"{'='*60}")
                    
                    loader = LLMDrivenExcelLoader(str(file_path))
                    
                    # 调用后端的process_batch方法，处理指定Sheet
                    with st.spinner(f"📊 正在处理 [{sheet_name}]..."):
                        # 修改load_data以支持指定工作表
                        original_load_data = loader.load_data
                        
                        def load_data_with_sheet():
                            return original_load_data(sheet_name=sheet_name)
                        
                        loader.load_data = load_data_with_sheet
                        
                        # 初始化新的处理器（避免结果混淆）
                        sheet_processor = BatchQuotationProcessor(
                            pricing_service=pricing_service,
                            sku_recommend_service=sku_service,
                            region=region_id
                        )
                        
                        results = sheet_processor.process_batch(loader, verbose=False)
                        
                        # 为每个结果添加Sheet来源标记
                        for result in results:
                            result['sheet_name'] = sheet_name
                        
                        all_results.extend(results)
                        
                        # 显示当前Sheet的统计
                        success_count = sum(1 for r in results if r.get('success', False))
                        logging.info(f"✅ [{sheet_name}] 处理完成: {success_count}/{len(results)} 成功")
                        
                        # 更新日志显示
                        log_container.code('\n'.join(log_handler.get_logs()), language='log')
                
                # 转换为DataFrame
                df_results = pd.DataFrame(all_results)
                
                logging.info(f"\n{'='*60}")
                logging.info("✅ 所有工作表处理完成！")
                logging.info(f"{'='*60}")
                
                # 统计信息（处理前）
                success_count = df_results['success'].sum() if 'success' in df_results.columns else 0
                total_count = len(df_results)
                
                # 计算总价（仅统计成功的记录）
                successful_df = df_results[df_results['success'] == True]
                total_price = successful_df['price_cny_month'].sum() if not successful_df.empty else 0
                
                # ============================================================
                # 格式化导出数据：调整列名、顺序和计算公式
                # ============================================================
                
                # 1. 列重命名映射
                column_mapping = {
                    'context_notes': '服务器类别',
                    'product_name': '产品名称',
                    'host_count': '服务数量',
                    'cpu_cores': 'CPU(core)',
                    'memory_gb': '内存(G)',
                    'storage_gb': '存储(G)',
                    'matched_sku': '产品规格',
                    'price_cny_month': '列表单价',
                    'workload_type': 'workload_type',
                    'success': 'success',
                    'error': 'error'
                }
                
                # 2. 重命名列
                df_export = df_results.rename(columns=column_mapping)
                
                # 3. 添加新列：折扣、折后总价
                df_export['折扣'] = ''  # 空白，用户手动填写
                df_export['折后总价'] = ''  # 初始为空，后续添加公式
                
                # 4. 选择并排序最终列（包含隐藏列）
                final_columns = [
                    '服务器类别', '产品名称', '服务数量', 
                    'success', 'error',  # 保留但稍后隐藏
                    'CPU(core)', '内存(G)', '存储(G)', 'workload_type',
                    '产品规格', '列表单价', '折扣', '折后总价'
                ]
                
                # 确保所有列都存在
                for col in final_columns:
                    if col not in df_export.columns:
                        df_export[col] = ''
                
                df_export = df_export[final_columns]
                
                logging.info(f"📊 总条数: {total_count}, 成功: {success_count}, 总价: ¥{total_price:,.2f}/月")
                
                # 最终更新日志显示
                log_container.code('\n'.join(log_handler.get_logs()), language='log')
                
                # 显示结果
                st.success("✅ 所有工作表处理完成！")
                
                st.markdown("---")
                st.subheader("📊 处理结果（汇总）")
                
                # 统计信息
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("总条数", total_count)
                with col2:
                    st.metric("成功", success_count)
                with col3:
                    st.metric("总价(月)", f"¥{total_price:,.2f}")
                
                # 显示详细结果表格（显示格式化后的数据）
                st.dataframe(df_export, use_container_width=True, height=400)
                
                # 导出Excel
                st.markdown("---")
                output_dir = Path("tests/output")
                output_dir.mkdir(parents=True, exist_ok=True)
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"quotation_{Path(uploaded_file.name).stem}_{timestamp}.xlsx"
                output_path = output_dir / output_filename
                
                # ============================================================
                # 使用openpyxl导出Excel，添加公式和隐藏列
                # ============================================================
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.styles import Font, Alignment
                
                wb = Workbook()
                ws = wb.active
                ws.title = "报价单"
                
                # 1. 写入表头
                headers = list(df_export.columns)
                ws.append(headers)
                
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # 2. 写入数据行（从第2行开始）
                for idx, row in df_export.iterrows():
                    row_data = []
                    for col in headers:
                        value = row[col]
                        # 处理NaN和空值
                        if pd.isna(value) or value == '':
                            row_data.append('')
                        else:
                            row_data.append(value)
                    ws.append(row_data)
                
                # 3. 添加折后总价公式（从第2行开始）
                # 公式: 列表单价 * 服务数量 * (1 - 折扣/100)
                # 找到各列的索引
                col_indices = {col: idx+1 for idx, col in enumerate(headers)}  # Excel列从1开始
                
                list_price_col = col_indices['列表单价']  # K列
                service_count_col = col_indices['服务数量']  # C列
                discount_col = col_indices['折扣']  # L列
                final_price_col = col_indices['折后总价']  # M列
                
                # 将数字转换为 Excel 列名 (1->A, 2->B, ...)
                def col_num_to_letter(n):
                    string = ""
                    while n > 0:
                        n, remainder = divmod(n - 1, 26)
                        string = chr(65 + remainder) + string
                    return string
                
                list_price_col_letter = col_num_to_letter(list_price_col)
                service_count_col_letter = col_num_to_letter(service_count_col)
                discount_col_letter = col_num_to_letter(discount_col)
                final_price_col_letter = col_num_to_letter(final_price_col)
                
                # 为每行添加公式（从第2行开始，因为第1行是表头）
                for row_idx in range(2, len(df_export) + 2):  # Excel行从1开始，数据从2开始
                    # 公式: =IF(L2="", K2*C2, K2*C2*(1-L2/100))
                    # 如果折扣为空，直接用单价*数量；否则计算折后价
                    formula = f'=IF({discount_col_letter}{row_idx}="", {list_price_col_letter}{row_idx}*{service_count_col_letter}{row_idx}, {list_price_col_letter}{row_idx}*{service_count_col_letter}{row_idx}*(1-{discount_col_letter}{row_idx}/100))'
                    ws[f'{final_price_col_letter}{row_idx}'] = formula
                
                # 4. 隐藏 success 和 error 列（设置列宽为0）
                success_col_idx = col_indices.get('success')
                error_col_idx = col_indices.get('error')
                
                if success_col_idx:
                    ws.column_dimensions[col_num_to_letter(success_col_idx)].hidden = True
                if error_col_idx:
                    ws.column_dimensions[col_num_to_letter(error_col_idx)].hidden = True
                
                # 5. 调整列宽（自动调整）
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # 跳过隐藏列
                    if ws.column_dimensions[column_letter].hidden:
                        continue
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 最大值50
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 保存Excel
                wb.save(output_path)
                
                with open(output_path, "rb") as f:
                    excel_data = f.read()
                
                st.download_button(
                    label="📥 下载Excel报价单",
                    data=excel_data,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # 清理临时文件
                try:
                    file_path.unlink()
                except:
                    pass
                
            except Exception as e:
                st.error(f"❌ 处理失败: {str(e)}")
                st.exception(e)

else:
    st.info("👆 请上传Excel文件开始处理")
