# -*- coding: utf-8 -*-
"""
Data Ingestion Layer - The Abstraction for Multimodal Input
Designed for extensibility: Today Excel, Tomorrow Images, Voice, etc.
"""
from dataclasses import dataclass
from typing import Iterator, Literal, Any
from abc import ABC, abstractmethod
import pandas as pd
from pathlib import Path


@dataclass
class QuotationRequest:
    """
    统一的报价请求数据结构
    
    这个抽象层是关键：无论输入来自Excel、图片、语音，都转换为此标准格式
    这样下游处理逻辑（Parser、Matcher、Pricer）完全不需要知道数据来源
    """
    source_id: str  # 数据来源标识 (e.g., "Row 1", "Screenshot_001.png")
    content: Any  # 主要内容 (文本/图片路径/音频路径等)
    content_type: Literal["text", "image", "audio"]  # 内容类型
    context_notes: str = ""  # 补充备注信息
    
    # 新增：直接从Excel提取的结构化数据
    product_name: str = "ECS"  # 产品名称 (默认ECS云服务器)
    host_count: int = 1  # 主机数
    cpu_cores: int = None  # CPU核心数
    memory_gb: int = None  # 内存(GB)
    storage_gb: int = None  # 存储(GB)
    
    def __str__(self) -> str:
        """便于调试的字符串表示"""
        content_preview = str(self.content)[:50] + "..." if len(str(self.content)) > 50 else str(self.content)
        return f"QuotationRequest(source={self.source_id}, type={self.content_type}, content={content_preview})"


class BaseDataLoader(ABC):
    """
    数据加载器抽象基类
    
    设计理念:
    - 定义统一的接口，让批处理器与具体数据格式解耦
    - 今天实现ExcelDataLoader，明天可以实现ImageDirLoader、VoiceTranscriptLoader
    - 批处理逻辑无需任何改动
    """
    
    @abstractmethod
    def load_data(self) -> Iterator[QuotationRequest]:
        """
        加载数据并转换为QuotationRequest流
        
        Yields:
            QuotationRequest: 标准化的报价请求对象
        """
        pass
    
    @abstractmethod
    def get_total_count(self) -> int:
        """
        获取总数据条数（用于进度显示）
        
        Returns:
            int: 数据总条数
        """
        pass


class ExcelDataLoader(BaseDataLoader):
    """
    Excel数据加载器 - BaseDataLoader的具体实现
    
    职责:
    - 读取Excel文件
    - 将每行数据转换为QuotationRequest对象
    - 处理列映射和数据清洗
    - 支持两种模式：简单模式（Specification列）和结构化模式（多列）
    """
    
    def __init__(self, file_path: str, spec_column: str = "Specification", remarks_column: str = "Remarks",
                 structured_mode: bool = False, skip_rows: int = 0):
        """
        初始化Excel加载器
        
        Args:
            file_path: Excel文件路径
            spec_column: 规格说明列名（简单模式）
            remarks_column: 备注列名（简单模式）
            structured_mode: 是否使用结构化模式（多列格式）
            skip_rows: 跳过前N行
        """
        self.file_path = Path(file_path)
        self.spec_column = spec_column
        self.remarks_column = remarks_column
        self.structured_mode = structured_mode
        self.skip_rows = skip_rows
        self._df = None
        
        # Validate file exists
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    def _load_dataframe(self) -> pd.DataFrame:
        """延迟加载DataFrame"""
        if self._df is None:
            try:
                # 使用openpyxl读取以更好地处理复杂Excel
                import openpyxl
                
                if self.structured_mode:
                    # 结构化模式：使用openpyxl直接读取
                    wb = openpyxl.load_workbook(self.file_path)
                    ws = wb.active
                    
                    # 提取所有行数据
                    rows_data = []
                    for row in ws.iter_rows(values_only=True):
                        rows_data.append(row)
                    
                    # 跳过指定行数
                    if self.skip_rows > 0:
                        rows_data = rows_data[self.skip_rows:]
                    
                    # 存储原始数据供后续处理
                    self._raw_rows = rows_data
                    self._df = pd.DataFrame()  # 结构化模式下不使用DataFrame
                else:
                    # 简单模式：使用pandas读取
                    self._df = pd.read_excel(self.file_path)
                    
                    # Validate required columns
                    if self.spec_column not in self._df.columns:
                        raise ValueError(f"Column '{self.spec_column}' not found in Excel. Available: {list(self._df.columns)}")
                    
                    # Remarks column is optional
                    if self.remarks_column not in self._df.columns:
                        print(f"⚠️  Warning: Column '{self.remarks_column}' not found. Using empty remarks.")
                        self._df[self.remarks_column] = ""
                
            except Exception as e:
                raise Exception(f"Failed to load Excel file: {e}")
        
        return self._df
    
    def load_data(self) -> Iterator[QuotationRequest]:
        """
        从Excel加载数据并转换为QuotationRequest流
        
        Yields:
            QuotationRequest: 每行数据对应一个请求对象
        """
        self._load_dataframe()
        
        if self.structured_mode:
            # 结构化模式：处理多列格式
            yield from self._load_structured_data()
        else:
            # 简单模式：处理Specification列
            yield from self._load_simple_data()
    
    def _load_simple_data(self) -> Iterator[QuotationRequest]:
        """简单模式：从Specification列加载数据"""
        df = self._df
        
        for idx, row in df.iterrows():
            # Extract specification content
            spec_content = str(row[self.spec_column]).strip()
            
            # Skip empty rows
            if not spec_content or spec_content.lower() in ['nan', 'none', '']:
                continue
            
            # Extract remarks (optional)
            remarks = str(row.get(self.remarks_column, "")).strip()
            if remarks.lower() in ['nan', 'none']:
                remarks = ""
            
            # Construct QuotationRequest
            yield QuotationRequest(
                source_id=f"Row {idx + 2}",  # Excel row number (1-indexed + header)
                content=spec_content,
                content_type="text",
                context_notes=remarks
            )
    
    def _load_structured_data(self) -> Iterator[QuotationRequest]:
        """
        结构化模式：从多列格式加载数据
        
        预期格式：
        第1行：标题
        第2行：主列名 (类型、服务器类别、安装内容、说明、主机数、虚拟机规格)
        第3行：CPU(核数)、内存(G)、数据盘(G)
        第4行及以后：数据行
        """
        if not hasattr(self, '_raw_rows') or not self._raw_rows:
            return
        
        # 找到包含CPU、内存、存储的列索引
        # 根据实际数据，这些列通常在第5-7列
        cpu_col_idx = 5  # CPU(核数)
        mem_col_idx = 6  # 内存(G)
        storage_col_idx = 7  # 数据盘(G)
        host_count_col_idx = 4  # 主机数
        desc_col_idx = 2  # 安装内容/说明
        
        # 从第4行开始读取数据（第1行是标题，第2-3行是表头）
        for row_idx, row in enumerate(self._raw_rows[3:], start=4):
            try:
                # 提取数据，确保不超过行长度
                if len(row) <= max(cpu_col_idx, mem_col_idx, storage_col_idx):
                    continue
                
                # 提取CPU、内存、存储
                cpu_value = row[cpu_col_idx] if cpu_col_idx < len(row) else None
                mem_value = row[mem_col_idx] if mem_col_idx < len(row) else None
                storage_value = row[storage_col_idx] if storage_col_idx < len(row) else None
                host_count_value = row[host_count_col_idx] if host_count_col_idx < len(row) else None
                desc_value = row[desc_col_idx] if desc_col_idx < len(row) else None
                
                # 跳过空行或非CPU数据行
                if cpu_value is None or mem_value is None:
                    continue
                
                # 转换为整数
                try:
                    cpu_cores = int(cpu_value) if cpu_value else None
                    memory_gb = int(mem_value) if mem_value else None
                    
                    # 存储可能是字符串或数字
                    if storage_value:
                        try:
                            storage_gb = int(storage_value)
                        except (ValueError, TypeError):
                            # 如果是字符串（如"500"），尝试转换
                            storage_str = str(storage_value).strip().replace(',', '')
                            storage_gb = int(storage_str) if storage_str.isdigit() else 0
                    else:
                        storage_gb = 0
                    
                    # 主机数
                    if host_count_value:
                        try:
                            # 处理"台"等单位
                            host_count_str = str(host_count_value).replace('台', '').strip()
                            host_count = int(host_count_str) if host_count_str.isdigit() else 1
                        except (ValueError, TypeError):
                            host_count = 1
                    else:
                        host_count = 1
                    
                    if cpu_cores is None or memory_gb is None:
                        continue
                    
                except (ValueError, TypeError):
                    # 无法转换为数字，跳过
                    continue
                
                # 构造描述文本
                desc_text = str(desc_value) if desc_value else ""
                content_text = f"{cpu_cores}C {memory_gb}G"
                if storage_gb > 0:
                    content_text += f" {storage_gb}G存储"
                if desc_text:
                    content_text += f" | {desc_text}"
                
                # 构造QuotationRequest
                yield QuotationRequest(
                    source_id=f"Row {row_idx + 1}",  # Excel原始行号
                    content=content_text,
                    content_type="text",
                    context_notes=desc_text,
                    host_count=host_count,
                    cpu_cores=cpu_cores,
                    memory_gb=memory_gb,
                    storage_gb=storage_gb
                )
                
            except Exception as e:
                # 跳过出错的行
                print(f"⚠️  Warning: Failed to parse row {row_idx + 1}: {e}")
                continue
    
    def get_total_count(self) -> int:
        """获取有效数据行数"""
        self._load_dataframe()
        
        if self.structured_mode:
            # 结构化模式：计算有CPU数据的行数
            if not hasattr(self, '_raw_rows') or not self._raw_rows:
                return 0
            
            count = 0
            cpu_col_idx = 5
            mem_col_idx = 6
            
            for row in self._raw_rows[3:]:  # 跳过表头
                if len(row) > max(cpu_col_idx, mem_col_idx):
                    cpu_value = row[cpu_col_idx]
                    mem_value = row[mem_col_idx]
                    if cpu_value is not None and mem_value is not None:
                        try:
                            int(cpu_value)
                            int(mem_value)
                            count += 1
                        except (ValueError, TypeError):
                            pass
            return count
        else:
            # 简单模式：计算非空行
            df = self._df
            valid_rows = df[self.spec_column].notna() & (df[self.spec_column] != "")
            return valid_rows.sum()


# ============================================================================
# LLM-Driven Excel Parser (Phase 7: Intelligent Adaptive Parsing)
# ============================================================================

class LLMDrivenExcelLoader(BaseDataLoader):
    """
    LLM驱动的Excel数据加载器 - 智能自适应解析
    
    核心理念:
    - 不依赖固定的列索引或表格结构
    - 读取整个Excel表格，提取所有有用信息
    - 使用Qwen-Plus LLM智能理解和结构化数据
    - 适应各种不同格式的报价单
    
    工作流程:
    1. 读取Excel原始数据（所有行列）
    2. 提取半结构化信息（数字、文本、位置关系）
    3. 构造Prompt提交给Qwen-Plus
    4. LLM返回标准化的JSON结构数据
    """
    
    def __init__(self, file_path: str, api_key: str = None):
        """
        初始化LLM驱动的Excel加载器
        
        Args:
            file_path: Excel文件路径
            api_key: DashScope API Key（如果不提供，从环境变量读取）
        """
        self.file_path = Path(file_path)
        self.api_key = api_key
        self._raw_rows = None
        self._parsed_data = None
        
        # Validate file exists
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Setup API key
        if not self.api_key:
            import os
            from dotenv import load_dotenv
            load_dotenv()
            self.api_key = os.getenv('DASHSCOPE_API_KEY')
            
        if not self.api_key:
            raise ValueError("DASHSCOPE_API_KEY is required for LLM-driven parsing")
    
    def _extract_semi_structured_data(self, sheet_name: str = None) -> str:
        """
        从Excel提取半结构化数据
        
        Args:
            sheet_name: 指定要读取的工作表名称，如果为None则读取活动工作表
        
        Returns:
            str: 格式化的半结构化文本，包含所有有用信息
        """
        import openpyxl
        
        # 使用 data_only=True 读取公式的计算结果，而不是公式本身
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 提取所有非空行数据
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            # 过滤掉完全空的行
            non_empty_cells = [cell for cell in row if cell is not None]
            if non_empty_cells:
                rows_data.append(row)
        
        self._raw_rows = rows_data
        
        # 构造半结构化文本表示
        semi_structured_text = "Excel表格数据：\n\n"
        
        for row_idx, row in enumerate(rows_data, 1):
            # 只保留非空单元格
            row_content = []
            for col_idx, cell in enumerate(row):
                if cell is not None:
                    # 识别数字和文本
                    cell_str = str(cell).strip()
                    if cell_str:
                        row_content.append(f"[列{col_idx+1}]{cell_str}")
            
            if row_content:
                semi_structured_text += f"第{row_idx}行: {' | '.join(row_content)}\n"
        
        return semi_structured_text
    
    def _parse_with_llm(self, semi_structured_data: str) -> list:
        """
        使用Qwen-Plus LLM解析半结构化数据
        
        Args:
            semi_structured_data: 半结构化文本数据
            
        Returns:
            list: 标准化的资源需求列表
        """
        from http import HTTPStatus
        import dashscope
        import json
        
        # 设置API Key
        dashscope.api_key = self.api_key
        
        # 构造Prompt
        system_prompt = """你是一个专业的云资源报价单解析专家。你的任务是从Excel表格数据中提取云服务器资源配置信息。

请仔细分析表格数据，识别出每一个资源配置项，并提取以下信息：
- 产品名称（如：ECS云服务器、PolarDB数据库、WAF防火墙、云安全中心等，默认为"ECS"）
- CPU核心数（整数）
- 内存大小GB（整数）
- 存储大小GB（整数，如果没有明确说明则为0）
- 主机数量（整数，默认1）
- 资源描述（简短描述这是什么服务或用途）

输出格式必须是JSON数组，每个元素包含：
{
  "row_number": 行号,
  "product_name": "产品名称",
  "cpu_cores": CPU核心数,
  "memory_gb": 内存GB,
  "storage_gb": 存储GB,
  "host_count": 主机数,
  "description": "资源描述"
}

注意事项：
1. 只提取实际的资源配置数据行，忽略标题、表头、空行
2. CPU和内存必须是有效的正整数
3. 产品名称识别规则：
   - 如果提到"数据库"、"MySQL"、"PolarDB"、"RDS" -> "PolarDB"
   - 如果提到"防火墙"、"WAF" -> "WAF"
   - 如果提到"安全"、"云安全中心"、"SAS" -> "云安全中心"
   - 其他情况默认为 "ECS"
4. 如果某行不包含资源配置信息，不要输出
5. 数字可能在不同的列中，需要智能识别
6. 主机数可能用"台"等单位，需要提取数字
7. 返回的JSON必须是有效的、可以直接解析的格式
"""
        
        user_prompt = f"""请分析以下Excel表格数据，提取所有云服务器资源配置信息：

{semi_structured_data}

请返回JSON数组格式的结果。"""
        
        try:
            # 调用Qwen-Plus
            response = dashscope.Generation.call(
                model='qwen-plus',
                messages=[
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user', 'content': user_prompt}
                ],
                result_format='message',
                temperature=0.1,  # 低温度，保证解析稳定性
            )
            
            if response.status_code == HTTPStatus.OK:
                llm_output = response.output.choices[0].message.content
                
                # 提取JSON（可能被markdown代码块包裹）
                json_str = llm_output.strip()
                if '```json' in json_str:
                    json_str = json_str.split('```json')[1].split('```')[0].strip()
                elif '```' in json_str:
                    json_str = json_str.split('```')[1].split('```')[0].strip()
                
                # 解析JSON
                parsed_data = json.loads(json_str)
                
                if not isinstance(parsed_data, list):
                    raise ValueError(f"LLM返回的不是数组格式: {type(parsed_data)}")
                
                return parsed_data
            else:
                raise Exception(f"LLM API调用失败: {response.code} - {response.message}")
                
        except json.JSONDecodeError as e:
            raise Exception(f"LLM返回的JSON格式无效: {e}\n原始输出: {llm_output}")
        except Exception as e:
            raise Exception(f"LLM解析失败: {e}")
    
    def load_data(self, sheet_name: str = None) -> Iterator[QuotationRequest]:
        """
        使用LLM智能加载和解析Excel数据
        
        Args:
            sheet_name: 指定要解析的工作表名称，如果为None则解析活动工作表
        
        Yields:
            QuotationRequest: 标准化的请求对象
        """
        # Step 1: 提取半结构化数据
        sheet_info = f" (工作表: {sheet_name})" if sheet_name else ""
        print(f"📖 正在读取Excel文件{sheet_info}...")
        semi_structured_data = self._extract_semi_structured_data(sheet_name)
        
        # Step 2: LLM智能解析
        print(f"🤖 正在使用Qwen-Plus智能解析表格{sheet_info}...")
        parsed_data = self._parse_with_llm(semi_structured_data)
        self._parsed_data = parsed_data
        
        print(f"✅ LLM成功解析出 {len(parsed_data)} 条资源配置{sheet_info}")
        
        # Step 3: 转换为QuotationRequest
        for idx, item in enumerate(parsed_data, 1):
            try:
                cpu_cores = int(item.get('cpu_cores', 0))
                memory_gb = int(item.get('memory_gb', 0))
                storage_gb = int(item.get('storage_gb', 0))
                host_count = int(item.get('host_count', 1))
                description = str(item.get('description', '')).strip()
                row_number = item.get('row_number', idx)
                product_name = str(item.get('product_name', 'ECS')).strip()
                
                # 验证必需字段
                if cpu_cores <= 0 or memory_gb <= 0:
                    print(f"⚠️  跳过无效配置[{idx}]: CPU={cpu_cores}, MEM={memory_gb}")
                    continue
                
                # 构造内容文本
                content_text = f"{cpu_cores}C {memory_gb}G"
                if storage_gb > 0:
                    content_text += f" {storage_gb}G存储"
                if description:
                    content_text += f" | {description}"
                
                # 在source_id中包含工作表信息
                sheet_prefix = f"{sheet_name} - " if sheet_name else ""
                yield QuotationRequest(
                    source_id=f"{sheet_prefix}Row {row_number} (LLM Parsed)",
                    content=content_text,
                    content_type="text",
                    context_notes=description,
                    product_name=product_name,
                    host_count=host_count,
                    cpu_cores=cpu_cores,
                    memory_gb=memory_gb,
                    storage_gb=storage_gb
                )
                
            except (ValueError, KeyError, TypeError) as e:
                print(f"⚠️  解析配置项[{idx}]失败: {e}")
                continue
    
    def get_total_count(self, sheet_name: str = None) -> int:
        """
        获取有效数据行数
        
        Args:
            sheet_name: 指定工作表名称
        
        如果已经解析过，返回解析结果数量；否则先解析
        """
        if self._parsed_data is not None:
            return len(self._parsed_data)
        
        # 执行一次完整解析来获取数量
        semi_structured_data = self._extract_semi_structured_data(sheet_name)
        parsed_data = self._parse_with_llm(semi_structured_data)
        self._parsed_data = parsed_data
        
        return len(parsed_data)


# ============================================================================
# Future Extension Point: Image-based Input (Placeholder)
# ============================================================================

class ImageDirLoader(BaseDataLoader):
    """
    图片目录加载器 (未来扩展)
    
    设计思路:
    - 遍历指定目录下的所有图片文件
    - 将图片路径封装为QuotationRequest
    - content_type设为"image"
    - 下游Parser检测到image类型后，调用Vision Model进行OCR/理解
    """
    
    def __init__(self, dir_path: str, supported_formats: tuple = ('.png', '.jpg', '.jpeg')):
        """
        初始化图片目录加载器
        
        Args:
            dir_path: 图片目录路径
            supported_formats: 支持的图片格式
        """
        self.dir_path = Path(dir_path)
        self.supported_formats = supported_formats
        
        if not self.dir_path.exists() or not self.dir_path.is_dir():
            raise ValueError(f"Invalid directory path: {dir_path}")
    
    def load_data(self) -> Iterator[QuotationRequest]:
        """
        从图片目录加载数据
        
        Yields:
            QuotationRequest: 图片路径封装的请求对象
        """
        image_files = [
            f for f in self.dir_path.iterdir()
            if f.suffix.lower() in self.supported_formats
        ]
        
        for img_file in sorted(image_files):
            yield QuotationRequest(
                source_id=img_file.name,
                content=str(img_file.absolute()),  # Full path to image
                content_type="image",
                context_notes=""  # Could be extracted from filename or metadata
            )
    
    def get_total_count(self) -> int:
        """获取图片文件总数"""
        return len([
            f for f in self.dir_path.iterdir()
            if f.suffix.lower() in self.supported_formats
        ])
